"""
Time: 2018.11.22
Author: Yuyuan's husband
title: Sentiment by SVM or Bayes
"""
from Tencent_word_Embedding import get_Word2Vec
from boson_dict import get_stop_word
from boson_dict import load_data
# from DQlihong_simple_word_embedding import load_Word2Vec_Model
from sklearn.cross_validation import train_test_split
from sklearn.preprocessing import scale
from sklearn.svm import SVC
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import SGDClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix
from sklearn import metrics
from sklearn.metrics import classification_report
from sklearn.naive_bayes import GaussianNB
from sklearn.naive_bayes import MultinomialNB
from sklearn.manifold import TSNE
from sklearn.externals import joblib
from xgboost import XGBClassifier
from openpyxl import load_workbook
import itertools
import jieba
import json
import numpy as np
import time
import matplotlib.pyplot as plt
from tqdm import tqdm
from boson_dict_config import bonson_config as bc

def get_tencent_word_embedding():
    # get tencent word embedding dict for convert phrase
    start = time.time()
    tencent_word_matrix = {}
    data = load_data()
    for uuid, txt in tqdm(data.items()):
        cut_txt = jieba.cut(txt[0], cut_all = False)
        for seg in cut_txt:
            if seg != '' and seg not in tencent_word_matrix:
                tencent_word_matrix[seg] = 0
        # break
    for i in get_Word2Vec(8824330):
        if i[0] in tencent_word_matrix:
            try:
                if len(i[1::]) == 200 and i[1::] != 0:
                    tencent_word_matrix[i[0]] = i[1::]
            except Exception as e:
                print ('Error type', e,i)
    with open('tencent_word_embedding_dict.json', 'w') as json_file:
        json.dump(tencent_word_matrix, json_file)
    json_file.close()
    end = time.time()
    # time almost equal 600s
    print ('JSON File has been done: ', end - start)

def get_data():
    pos_list = []
    neg_list = []
    neutral_list = []
    data = load_data()
    for uuid in data:
        if data[uuid][1] == 0:
            neg_list.append(data[uuid][0])
        elif data[uuid][1] == 1:
            neutral_list.append(data[uuid][0])
        else:
            pos_list.append(data[uuid][0])
    # try to add train data to build balance train data set
    pos_list = pos_list*6
    neg_list = neg_list*2
    # print ('<===pos_list, neutral_list, neg_list,pos_list===>', neutral_list, neg_list)
    return pos_list, neutral_list, neg_list

def prepare_train_data(train_X, y, ratio):
    train_x, test_x, train_y, test_y = train_test_split(train_X, y, test_size = ratio)
    return train_x, test_x, train_y, test_y

def shuttle_data():
    pass

def convert_Embedding(sentence_list, stop_word, model):
    """
    for each sentence, the mean vector of all its vectors is used to represent this sentence
    :param pos_list:[txt,...]
    :param neg_list:[txt,...]
    :param neutral_list:[txt,...]
    :param model:loaded model by word2vec
    :return:pos_input、neutral_input、neg_input
    """
    res_input = []
    for sentence in sentence_list:
        word_input = []
        seg_list = jieba.cut(sentence, cut_all = False)
        for seg in seg_list:
            if seg not in stop_word and seg != '':
                try :
                    word_input.append(model[seg])
                except KeyError:
                    continue
        res_list = np.array(word_input, dtype = float)
        if len(res_list) != 0:
            result_array = sum(np.array(res_list))/len(res_list)
            res_input.append(result_array)
    # print (res_input)
    return res_input

def build_vecs_tencent(pos_list, neg_list, neutral_list):
    json_dict = json.loads(open('tencent_word_embedding_dict.json', 'r', encoding = 'utf-8').read())
    def convert_tencent_embedding(sentence):
        res_input = []
        for sen in sentence:
            word_input = []
            seg_list = jieba.cut(sen, cut_all= False)
            for seg in seg_list:
                try:
                    if seg in json_dict.keys() and json_dict[seg] != 0:
                        word_input.append(np.array(json_dict[seg]))
                except KeyError:
                    continue
            # print (word_input)
            res_list = np.array(word_input, dtype=float)
            if len(res_list) != 0:
                result_array = sum(np.array(res_list)) / len(res_list)
                res_input.append(result_array)
        return res_input
    pos_input = convert_tencent_embedding(pos_list)
    neg_input = convert_tencent_embedding(neg_list)
    neutral_input = convert_tencent_embedding(neutral_list)
    print('RES_len:', len(pos_input), len(neg_input), len(neutral_input))
    y = np.concatenate((np.ones(len(pos_input)) * 2, np.zeros(len(neg_input)), np.ones(len(neutral_input))))
    X = pos_input[:]
    for neg in neg_input:
        X.append(neg)
    for neu in neutral_input:
        X.append(neu)
    train_X = np.array(X)
    # save_list_x = train_X.tolist()
    # save_list_y = y.tolist()
    # with open('train_x.json', 'w') as json_file:
    #     json.dump(save_list_x, json_file)
    # with open('train_y.json', 'w') as json_file:
    #     json.dump(save_list_y, json_file)
    return train_X, y

def build_vecs(pos_list, neg_list, neutral_list, stop_word,model):
    pos_input = convert_Embedding(pos_list,stop_word,model)
    neg_input = convert_Embedding(neg_list,stop_word,model)
    neutral_input = convert_Embedding(neutral_list,stop_word,model)
    # print(len(pos_list), len(neg_list),len(neutral_list))
    # use 2 for positive sentiment, 0 for negative sentiment, 1 for neutral sentiment
    y = np.concatenate((np.ones(len(pos_input))*2,np.zeros(len(neg_input)),np.ones(len(neutral_input))))
    X = pos_input[:]
    for neg in neg_input:
        X.append(neg)
    for neu in neutral_input:
        X.append(neu)
    train_X = np.array(X)
    # print (train_X.shape, y.shape, len(pos_input), len(neg_input), len(neutral_input))
    return train_X, y

def plt_with_tsne(train_x, pos_list, neg_list, neutral_list):
    ts = TSNE(2)
    reduced_vec = ts.fit_transform(train_x)
    for i in range(len(reduced_vec)):
        if i < len(pos_list):
            color = 'b'
        elif i >= len(pos_list) and i < len(pos_list) + len(neg_list):
            color = 'r'
        else:
            color = 'g'
        plt.plot(reduced_vec[i, 0], reduced_vec[i, 1], marker = 'o', color = color, markersize = 8)
    plt.show()

def plot_pca(train_x):
    #type error in pca.fit
    pca = PCA(n_components = 200)
    pca.fit(train_x)
    plt.figure(1, figsize = (4,3))
    plt.clf()
    plt.axes([.2,.2,.7,.7])
    plt.plot(pca.explained_variance_, linewidth = 3)
    plt.axis('tight')
    plt.xlabel('n_components')
    plt.ylabel('explained_variance')
    plt.show()
    reduced_x = PCA(n_components = 50).fit_transform(train_x)
    # print each explained_variance_tatio
    # print(pca.explained_variance_ratio_)
    return reduced_x

def RandomForest_Model(train_x, test_x, train_y, test_y):

    # RFC = RandomForestClassifier(max_features = 200, min_samples_leaf=50, n_estimators=400)
    RFC = RandomForestClassifier(min_samples_leaf=3, n_estimators=400)
    RFC_M = RFC.fit(train_x, train_y)
    print ('RandomForest Accuarcy: ', RFC_M.score(test_x, test_y))
    pred_y = RFC.fit(train_x, train_y).predict(test_x)
    classify_report = metrics.classification_report(test_y, pred_y)
    print('classify_report', classify_report)
    # pred_y = RFC.fit(train_x, train_y).predict(test_x)
    cnf_matrix = confusion_matrix(test_y, pred_y)
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='RandomForest Confusion matrix')
    plt.show()
    # joblib.dump(RFC, 'RFC_modify_train_data_model.pkl')

def sgd_model(train_x, test_x, train_y, test_y):
    lr = SGDClassifier(loss = 'log', penalty = 'l1')
    lr.fit(train_x, train_y)
    print ('sgd test Accuracy: %.2f'%lr.score(test_x, test_y))
    pred_y = lr.fit(train_x, train_y).predict(test_x)
    classify_report = metrics.classification_report(test_y, pred_y)
    print('classify_report', classify_report)
    # pred_y = lr.fit(train_x, train_y).predict(test_x)
    cnf_matrix = confusion_matrix(test_y, pred_y)
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='SGD Confusion matrix')
    plt.show()

def svm_model(train_x, test_x, train_y, test_y):
    # tencent word embedding test accuarcy 0.6
    # clf = SVC(C=10, kernel='rbf', degree=3, gamma=0.01, coef0=.1, probability=True)
    clf = SVC(C=10, kernel='rbf', degree=3, gamma=0.01, probability=True)
    clf.fit(train_x, train_y)
    pred_y = clf.fit(train_x, train_y).predict(test_x)
    print('RBF svm test Accuracy: %.2f' % clf.score(test_x, test_y))
    classify_report = metrics.classification_report(test_y, pred_y)
    print('classify_report', classify_report)
    ### some problem here
    # pred_y = clf.fit(train_x, train_y).predict(test_x)
    # print (classification_report(test_y, pred_y))
    cnf_matrix = confusion_matrix(test_y, pred_y)
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='SVM Confusion matrix')
    plt.show()
    # save model
    # joblib.dump(clf, 'RBF_SVM_modify_train_data_model.pkl')

def one_hot_naive_bayes_model(stopword_dir):
    pos_list, neutral_list, neg_list = get_data()
    # print (pos_list[0])
    res_dict = {}
    stop_word_list = get_stop_word(stopword_dir)
    def get_dict(list_, j):
        for sen in list_:
            for word in sen:
                if word != '' and word not in res_dict.keys() and word not in stop_word_list:
                    res_dict[word] = j
                    j += 1
        return j
    num = get_dict(pos_list, j=0)
    num2 = get_dict(neg_list, num)
    num3 = get_dict(neutral_list, num2)
    id_to_char = {value: key for key, value in res_dict.items()}
    max_features = 150
    def get_ont_hot(list_):
        final_list = []
        for sen in list_:
            input_list = []
            for i in range(max_features):
                try:
                    if sen[i] in res_dict:
                        input_list.append(res_dict[sen[i]])
                    else:
                        input_list.append(5000)
                except Exception as e:
                    padding = 5000
                    input_list.append(padding)
            # print (input_list)
            # break
            res_list = np.array(input_list)
            # print (res_list)
            final_list.append(res_list)
        return final_list
    pos_input_list = get_ont_hot(pos_list)
    neutral_input_list = get_ont_hot(neutral_list)
    neg_input_list = get_ont_hot(neg_list)
    y = np.concatenate(
        (np.ones(len(pos_input_list)) * 2, np.zeros(len(neg_input_list)), np.ones(len(neutral_input_list))))
    X = pos_input_list[:]
    for neg in neg_input_list:
        X.append(neg)
    for neu in neutral_input_list:
        X.append(neu)
    train_X = np.array(X)
    # print (len(train_X))
    # print (train_X[0])
    train_x, test_x, train_y, test_y = train_test_split(train_X, y, test_size=0.2)
    mnb = MultinomialNB()
    y_pred = mnb.fit(train_x, train_y).predict(test_x)
    # right_num = (test_y == y_pred).sum()
    # print("GaussianNB naive bayes accuracy :%f" % (float(right_num) / test_x.shape[0]))
    print('The accuracy of Navie Bayes Classifier is', mnb.score(test_x, test_y))
    # MNB = MultinomialNB(alpha=0.000607)
    # raise error Input X must be non-negative, that means that marixs
    # y_pred = MNB.fit(train_x, train_y).predict(test_x)
    # right_num = (test_y == y_pred).sum()
    # print("MultinomialNB naive bayes accuracy :%f" % (float(right_num) / test_x.shape[0]))
    cnf_matrix = confusion_matrix(test_y, y_pred)
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='one_hot Naive Bayes Confusion matrix')
    plt.show()

def naive_model(train_x, test_x, train_y, test_y):
    gnb = GaussianNB()
    # mnb = MultinomialNB()
    y_pred = gnb.fit(train_x, train_y).predict(test_x)
    # right_num = (test_y == y_pred).sum()
    # print("GaussianNB naive bayes accuracy :%f" % (float(right_num) / test_x.shape[0]))
    print('The accuracy of Navie Bayes Classifier is', gnb.score(test_x, test_y))
    classify_report = metrics.classification_report(test_y, y_pred)
    print('classify_report', classify_report)
    # MNB = MultinomialNB(alpha=0.000607)
    # raise error Input X must be non-negative, that means that marixs
    # y_pred = MNB.fit(train_x, train_y).predict(test_x)
    # right_num = (test_y == y_pred).sum()
    # print("MultinomialNB naive bayes accuracy :%f" % (float(right_num) / test_x.shape[0]))
    cnf_matrix = confusion_matrix(test_y, y_pred)
    # print("Precision", metrics.precision_score(test_y, y_pred))
    # print("Recall", metrics.recall_score(test_y, y_pred))
    # print("f1_score", metrics.f1_score(test_y, y_pred))
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='Naive Bayes Confusion matrix')
    plt.show()

def logistic_model(train_x, test_x, train_y, test_y):
    LR_model = LogisticRegression(penalty = 'l1', C= 1, solver = 'liblinear',
                                  multi_class = 'ovr')
    LR_model = LR_model.fit(train_x, train_y)
    y_pred = LR_model.predict(test_x)
    cnf_matrix = confusion_matrix(test_y, y_pred)
    # modify recall and accuracy format if u want double classicfy
    print("logistic accuracy metric in the testing dataset: ",LR_model.score(test_x, test_y))
    classify_report = metrics.classification_report(test_y, y_pred)
    print('classify_report',classify_report)
    # print("Precision", metrics.precision_score(test_y, y_pred))
    # print("Recall", metrics.recall_score(test_y, y_pred))
    # print("f1_score", metrics.f1_score(test_y, y_pred))
    # print("logistic Recall metric in the testing dataset: ", cnf_matrix[2, 2] / (cnf_matrix[2, 0] + cnf_matrix[2, 1] + cnf_matrix[2,2]))
    # print("logistic accuracy metric in the testing dataset: ", (cnf_matrix[1, 1] + cnf_matrix[0, 0] + cnf_matrix[2, 2])/ (
    #     cnf_matrix[0, 0] + cnf_matrix[1, 1] + cnf_matrix[1, 0] + cnf_matrix[0, 1] + cnf_matrix[0, 2] + cnf_matrix[1, 2]+
    #     cnf_matrix[2, 2] + cnf_matrix[2, 0] + cnf_matrix[2, 1] +cnf_matrix[2, 2]))  # Plot non-normalized confusion matrix
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='Logistic Confusion matrix')
    plt.show()

def xgb_model(train_x, test_x, train_y, test_y):
    xgbc = XGBClassifier(max_depth = 10,early_stopping_rounds = True,n_estimators=400,colsample_bytree = 0.8,min_child_weight = 5)
    xgb = xgbc.fit(train_x, train_y)
    pred = xgb.predict(test_x)
    print('XGBoost test Accuracy: %.2f' % xgb.score(test_x, test_y))
    classify_report = metrics.classification_report(test_y, pred)
    print('classify_report', classify_report)
    cnf_matrix = confusion_matrix(test_y, pred)
    class_names = [0, 1, 2]
    plt.figure()
    plot_confusion_matrix(cnf_matrix, classes=class_names, title='XGBoost Confusion matrix')
    plt.show()
    # save model
    # joblib.dump(xgb, 'xgboost_modify_train_data_model.pkl')

def Snow_naive_bayes(data_dir):
    from snownlp import SnowNLP
    wb = load_workbook(data_dir)
    ws = wb.get_sheet_by_name('Sheet1')
    res_dict = {}
    for i in range(1, ws.max_row):
        single = ws.cell(row=i + 1, column=9).value
        txt = ws.cell(row=i + 1, column=4).value
        res_dict[txt] = single
    com_dict = {}
    for key in res_dict:
        sen = SnowNLP(key).sentiments
        if sen < 0.4:
            sin = 0
        elif sen < 0.6:
            sin = 1
        else:
            sin = 2
        com_dict[key] = sin
    j = 0
    for key, value in com_dict.items():
        if res_dict[key] != value:
            j += 1
    print("SnowNLP model Accuracy: ", j / len(res_dict))

def plot_confusion_matrix(cm, classes, title='Confusion matrix', cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    """
    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=0)
    plt.yticks(tick_marks, classes)
    thresh = cm.max() / 2.
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        plt.text(j, i, cm[i, j],
                 horizontalalignment="center",
                 color="white" if cm[i, j] > thresh else "black")
    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('Predicted label')

def get_sentiment_model(data_dir):
    start = time.time()
    # stop_word = get_stop_word()
    # use tencent word embedding
    pos_list, neutral_list, neg_list = get_data()
    train_X, y = build_vecs_tencent(pos_list, neutral_list, neg_list)
    # model = load_Word2Vec_Model()
    # train_X, y = build_vecs(pos_list, neutral_list, neg_list, stop_word, model)
    # show the data distribute...it costs too long...
    # plt_with_tsne(train_X, pos_list, neg_list, neutral_list)
    train_ratio = 0.2
    # standardization , tencent word embedding has been done.
    # train_X = scale(train_X)
    # plot the PCA spectrum
    # reduce or not is no necessary
    # reduce_x = plot_pca(train_X)
    # reduce_train_x, reduce_test_x, reduce_train_y, reduce_test_y = prepare_train_data(reduce_x, y, train_ratio)
    train_x, test_x, train_y, test_y = prepare_train_data(train_X, y, train_ratio)
    # use RandomForest to classify accuracy equal
    RandomForest_Model(train_x, test_x, train_y, test_y)
    # use SGD of logistic to classify accuracy equal 0.53
    sgd_model(train_x, test_x, train_y, test_y)
    # sgd_model(reduce_train_x, reduce_test_x, reduce_train_y, reduce_test_y)
    # use RBF svm to classify accuracy equal 0.50
    svm_model(train_x, test_x, train_y, test_y)
    # svm_model(reduce_train_x, reduce_test_x, reduce_train_y, reduce_test_y)
    # use naive bayes to classify accuracy equal 0.51
    naive_model(train_x, test_x, train_y, test_y)
    # use logistic to classify accuracy equal 0.57
    logistic_model(train_x, test_x, train_y, test_y)
    xgb_model(train_x, test_x, train_y, test_y)
    Snow_naive_bayes(data_dir)
    end = time.time()
    print ('Total Time :', end - start)

# def get_test_data():
#     start = time.time()
#     wb = load_workbook(r'C:\Dissertation\微博测试样例数据-1113.xlsx')
#     ws = wb.get_sheet_by_name('Sheet1')
#     # print (ws.cell(row= 2, column=2).value)
#     test_list = []
#     for i in range(1, ws.max_row):
#         test_list.append(ws.cell(row = i+1, column = 2).value)
#     load_tencent_dict = json.loads(open('tencent_word_embedding_dict.json', 'r').read())
#     new_word = {}
#     for txt in test_list:
#         seg_list = jieba.cut(txt, cut_all = False)
#         for seg in seg_list:
#             if seg not in load_tencent_dict and seg != '':
#                 new_word[seg] = 0
#     # 787 new seg word has been already get
#     # print (len(new_word))
#     for i in get_Word2Vec(8824330):
#         if i[0] in new_word:
#             try:
#                 if len(i[1::]) == 200 and i[1::] != 0:
#                     new_word[i[0]] = i[1::]
#             except Exception as e:
#                 print ('Error type', e,i)
#     merge_dict = dict(new_word, **load_tencent_dict)
#     with open('tencent_word_embedding_dict.json', 'w+') as json_file:
#         json.dump(merge_dict, json_file)
#     json_file.close()
#     end = time.time()
#     print ('JSON File has been done: ', end - start)

def _test_save_model_validate():
    # when use new test only need load once
    # get_test_data()
    def get_new_test_list():
        pos_list = []
        neg_list = []
        neutral_list = []
        wb = load_workbook(r'test_data_dir')
        ws = wb.get_sheet_by_name('Sheet1')
        for i in range(1, ws.max_row):
            single = ws.cell(row = i+1, column = 4).value
            txt = ws.cell(row = i+1, column = 2).value
            if single == 0:
                neg_list.append(txt)
            elif single == 1:
                neutral_list.append(txt)
            else:
                pos_list.append(txt)
        return pos_list, neutral_list, neg_list

    pos_list, neutral_list, neg_list = get_new_test_list()
    print ('Test_Data has been load')
    def loading_save_model(model, test_x, test_y, classifyname):
        y_pred = model.predict(test_x)
        cnf_matrix = confusion_matrix(test_y, y_pred)
        class_names = [0, 1, 2]
        plt.figure()
        plot_confusion_matrix(cnf_matrix, classes=class_names, title='%s Confusion matrix'%classifyname)
        plt.show()
        print('%s test Accuracy: %.2f' %(classifyname, model.score(test_x, test_y)))

    test_X, test_Y = build_vecs_tencent(pos_list, neutral_list, neg_list)
    svm = joblib.load('RBF_SVM_modify_train_data_model.pkl')
    loading_save_model(svm, test_X, test_Y, 'SVM')
    xgb = joblib.load('xgboost_modify_train_data_model.pkl')
    loading_save_model(xgb, test_X, test_Y, 'XGBoost')
    clf = joblib.load('RFC_modify_train_data_model.pkl')
    loading_save_model(clf, test_X, test_Y, 'RFC')

def real_predict(test_x):
    json_dict = json.loads(open('tencent_word_embedding_dict.json', 'r', encoding='utf-8').read())
    seg_list = jieba.cut(test_x, cut_all = False)
    test_X = []
    res_input = []
    for seg in seg_list:
        try:
            if seg in json_dict.keys() and json_dict[seg] != 0:
                test_X.append(np.array(json_dict[seg]))
        except KeyError:
            continue
    res_list = np.array(test_X, dtype=float)
    if len(res_list) != 0:
        result_array = sum(np.array(res_list)) / len(res_list)
    # print ('result_array',result_array)
    res_input.append(result_array)
    # print ('res_input',res_input)
    test = np.array(res_input)
    # print ('test',test)
    def load_model_predict(model, test, classifyname):
        y_pred = model.predict(test)
        # print (y_pred[0], type(y_pred))
        print ('%s Model predict result: %.2f'%(classifyname, y_pred[0]))
    svm = joblib.load('RBF_SVM_modify_train_data_model.pkl')
    load_model_predict(svm, test, 'SVM')
    xgb = joblib.load('xgboost_modify_train_data_model.pkl')
    load_model_predict(xgb, test, 'XGBoost')
    clf = joblib.load('RFC_modify_train_data_model.pkl')
    load_model_predict(clf, test, 'RFC')

def get_variable_model_test():
    # get regular train and test data
    # pos_list, neutral_list, neg_list = get_data()
    # train_X, y = build_vecs_tencent(pos_list, neutral_list, neg_list)
    # train_x, test_x, train_y, test_y = train_test_split(train_X, y, test_size = 0.3)
    # train_x = train_x.tolist()
    # test_x = test_x.tolist()
    # train_y = train_y.tolist()
    # test_y = test_y.tolist()
    # json.dump([train_x, test_x, train_y, test_y],open('train_x.json', 'w'))
    train_x, test_x, train_y, test_y = json.load(open('train_x.json', 'r'))
    train_x = np.array(train_x)
    test_x = np.array(test_x)
    train_y = np.array(train_y)
    test_y = np.array(test_y)
    # sgd_model(train_x, test_x, train_y, test_y)
    # logistic_model(train_x, test_x, train_y, test_y)
    # naive_model(train_x, test_x, train_y, test_y)
    # one_hot_naive_bayes_model()
    # RandomForest_Model(train_x, test_x, train_y, test_y)
    svm_model(train_x, test_x, train_y, test_y)
    # xgb_model(train_x, test_x, train_y, test_y)

if __name__ == "__main__":

    get_tencent_word_embedding()
    import random
    random.seed(1)
    get_sentiment_model(bc.data)
    _test_save_model_validate()
    test_x = '新买的手机的信号还可以，没有预期的期望高，不过对于我来说也还可以了'
    real_predict(test_x)
    get_variable_model_test()