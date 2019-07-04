"""
time:2018.11.15
author:Yuyuan's husband
title: Boson_Emotional_Dict_Method
"""
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import jieba
import time
import matplotlib.pyplot as plt

def load_jieba(compile_sentiment_dict):
    """
    add compile_sentiment_dict into jieba
    :param compile_sentiment_dict: {'compile phrase':score}
    :return: add compile words into jieba dictionary
    """
    for words in compile_sentiment_dict:
        jieba.add_word(words)

def get_sentiment_dict(boson_dict):
    """
    get sentiment dict and compile dict
    :param boson_dict: {'Words':'Boson_Score'}
    :return: {'compile phrase':score},{'pos':[...],'neg':[...]}
    """
    sentiment_dict = {}
    re_sentiment_dict = {}
    # print(boson_dict.keys())
    file = open(r'C:\Dissertation\sentiment\正面评价词语（中文）1.txt', 'r', encoding='utf-8').read()
    file2 = open(r'C:\Dissertation\sentiment\正面情感词语（中文）1.txt', 'r', encoding='utf-8').read()
    file3 = open(r'C:\Dissertation\sentiment\负面评价词语（中文）1.txt', 'r', encoding='utf-8').read()
    file4 = open(r'C:\Dissertation\sentiment\负面情感词语（中文）1.txt', 'r', encoding='utf-8').read()
    sentiment_dict['pos'] = [word.strip() for word in file.split('\n')[1::] if word != '']
    sentiment_dict['neg'] = [word.strip() for word in file3.split('\n')[1::] if word != '']
    for word in file2.split('\n')[1::]:
        if word != '':
            sentiment_dict['pos'].append(word.strip())
    for word in file4.split('\n')[1::]:
        if word != '':
            sentiment_dict['neg'].append(word.strip())
    # print ("=========>", sentiment_dict.keys())
    # print ("sentiment_dict.keys() ",sentiment_dict.keys())
    for flag in sentiment_dict:
        for phrase in sentiment_dict[flag]:
            if phrase in boson_dict.keys():
                re_sentiment_dict[phrase] = boson_dict[phrase]
    # print('result:',re_sentiment_dict)
    return re_sentiment_dict,sentiment_dict

def get_adverb_dict():
    """
    get adverb_dict
    :return: {'extreme':[...],...,'insufficently':[...]}
    """
    adverb_dict = {}
    adverb_dict['extreme'] = ['百分之百', '倍加', '备至', '不得了', '不堪', '不可开交', '不亦乐乎', '不折不扣',
                            '彻头彻尾', '充分', '到头', '地地道道', '非常', '极', '极度', '极端', '极其', '极为',
                            '截然', '尽', '惊人地', '绝', '绝顶', '绝对', '绝对化', '刻骨', '酷', '满', '满贯',
                            '满心', '莫大', '奇', '入骨', '甚为', '十二分', '十分', '十足', '死', '滔天', '痛',
                            '透', '完全', '完完全全', '万', '万般', '万分', '万万', '无比', '无度', '无可估量',
                            '无以复加', '无以伦比', '要命', '要死', '已极', '已甚', '异常', '逾常', '贼', '之极',
                            '之至', '至极', '卓绝', '最为', '佼佼', '郅', '綦', '齁', '最']
    adverb_dict['very'] = ['不过', '不少', '不胜', '惨', '沉', '沉沉', '出奇', '大为', '多', '多多', '多加',
                         '多么', '分外', '格外', '够瞧的', '够戗', '好', '好不', '何等', '很', '很是',
                         '坏', '可', '老', '老大', '良', '颇', '颇为', '甚', '实在', '太', '太甚', '特', '特别',
                         '尤', '尤其', '尤为', '尤以', '远', '着实', '曷', '碜']
    adverb_dict['more'] = [ '大不了', '多', '更', '更加', '更进一步', '更为', '还', '还要', '较', '较比',
                          '较为', '进一步', '那般', '那么', '那样', '强', '如斯', '益', '益发', '尤甚',
                          '逾', '愈', '愈发', '愈加', '愈来愈', '愈益', '远远', '越发', '越加', '越来越',
                          '越是', '这般', '这样', '足', '足足']
    adverb_dict['ish'] = ['点点滴滴', '多多少少', '怪', '好生', '还', '或多或少', '略', '略加', '略略', '略微', '略为',
                        '蛮', '稍', '稍稍', '稍微', '稍为', '稍许', '挺', '未免', '相当', '些', '些微', '些小',
                        '一点', '一点儿', '一些', '有点', '有点儿', '有些']
    adverb_dict['insufficiently'] = ['半点', '不大', '不丁点儿', '不甚', '不怎么',
                                   '聊', '没怎么', '轻度', '弱', '丝毫', '微', '相对']
    return adverb_dict

def get_privative_list():
    """
    get privative_list
    :return: [...]
    """
    file = open(r'C:\Dissertation\sentiment\否定词.txt', 'r', encoding='utf-8').read()
    stop_privative_list = file.split('\n')
    return stop_privative_list

def load_data():
    """
    get source weibo data
    :return: {'uuid':['text', '0 or 1 or 2']}
    """
    wb = load_workbook(r'C:\Dissertation\全网筛选数据-联通-汇总-1113.xlsx')
    ws = wb.get_sheet_by_name('微博')
    data_dict = {}
    for i in range(1, ws.max_row):
        try:
            weibo_score = [ws.cell(row = i+1, column = 4).value, ws.cell(row = i+1, column = 9).value]
            data_dict[ws.cell(row = i+1, column = 1).value] = weibo_score
        except KeyError:
            print ("KeyError_Line:: ", i+1)
            continue
    return data_dict

def Boson_dict():
    """
    get Boson_Emotional_dict
    :return: {'Words':'Boson_Score'}
    """
    Boson_file = open(r'C:\Dissertation\BosonNLP_sentiment_score\BosonNLP_sentiment_score\BosonNLP_sentiment_score.txt', 'r',
                encoding='utf-8').read()
    Boson_dict = {}
    Boson_file = Boson_file.split('\n')
    for boson in Boson_file:
        couple = boson.split(' ')
        try:
            Boson_dict[couple[0]] = round(float(couple[1]),4)
        except IndexError:
            continue
    return Boson_dict

def get_stop_word():
    """
    get Stop_word_list
    :return: [...]
    """
    file = open(r'C:\Dissertation\stop.txt', 'r', encoding='utf-8').read()
    stop_word_list = file.split('\n')
    return stop_word_list

def get_Boson_score(weibo, Boson_dict, stop_word_list, compile_sentiment_dict, sentiment_dict, adverb_dict, privative_list):
    """
    get Boson_Model_Emotion_Score
    Modify by Emotional_Dictionary,but where is it?
    :param weibo:{'uuid':['text', '0 or 1 or 2']}
    :param Boson_dict:{'Words':'Boson_Score'}
    :param stop_word_list:[...]
    :param compile_sentiment_dict:{'compile phrase':score}
    :param sentiment_dict:{'pos':[...],'neg':[...]}
    :param adverb_dict:{'extreme':[...],...,'insufficently':[...]}
    :param privative_list:[...]
    :return:float(score)
    """
    seg_list = jieba.cut(weibo, cut_all = False)
    seg_to_stop_list = [seg for seg in seg_list if seg not in stop_word_list]
    re_privative_list = [seg for seg in seg_to_stop_list if seg in privative_list]
    re_compile_sentiment_list = [[seg,compile_sentiment_dict[seg]] for seg in seg_to_stop_list if seg in compile_sentiment_dict.keys()]
    re_adverb_list = [seg for seg in seg_to_stop_list for key in adverb_dict if seg in adverb_dict[key]]
    Emotional_Score = 0
    weight = 1
    for word in seg_to_stop_list:
        if word in re_privative_list:
            weight = -1*weight
        elif word in re_compile_sentiment_list:
            Emotional_Score += re_compile_sentiment_list[word][1]
        elif word in re_adverb_list:
            if word in adverb_dict['extreme']:
                Emotional_Score += 2*Emotional_Score
            if word in adverb_dict['very']:
                Emotional_Score += 1.75*Emotional_Score
            if word in adverb_dict['more']:
                Emotional_Score += 1.5*Emotional_Score
            if word in adverb_dict['ish']:
                Emotional_Score += 1.3*Emotional_Score
            if word in adverb_dict['insufficiently']:
                Emotional_Score += 1.1*Emotional_Score
        else:
            for key in sentiment_dict:
                if word in sentiment_dict[key]:
                    if word in Boson_dict:
                        if Boson_dict[word] < 0 and key == 'neg':
                            Emotional_Score += Boson_dict[word]
                        elif Boson_dict[word] > 0 and key == 'pos':
                            Emotional_Score += Boson_dict[word]
                        else:
                            Emotional_Score += -1*Boson_dict[word]
    # try:
    #     Emotional_Score = round(Emotional_Score/no_neutral, 4)
    # except ZeroDivisionError:
    #     Emotional_Score = 0
    return Emotional_Score

def Accuracy_Boson_Model(data_dict, res_dict):
    """
    Accuracy_Boson_Model
    :param data_dict:{'uuid':['text', '0 or 1 or 2']}
    :param res_dict:{uuid:score,...}
    :return: several accuracy score
    """
    res = 0
    neutral = 0
    res_neu = 0
    pos = 0
    res_pos = 0
    neg = 0
    res_neg = 0
    for uuid in data_dict:
        res_score = res_dict[uuid]
        org_score = int(data_dict[uuid][1])
        if org_score == 1:
            neutral += 1
            if res_score == org_score:
                res_neu += 1
        if org_score == 0:
            neg += 1
            if res_score == org_score:
                res_neg += 1
        if org_score == 2:
            pos += 1
            if res_score == org_score:
                res_pos += 1
        if res_score == org_score:
            res += 1
    ratio_total= res/len(data_dict)
    ratio_pos = res_pos/pos
    ratio_neu = res_neu/neutral
    ratio_neg = res_neg/neg
    # print ('res_neg, neg:', res_neg, neg)
    return ratio_total, ratio_pos, ratio_neu, ratio_neg

def boson_Dict_Model(data_dict, Boson_dict, compile_sentiment_dict, sentiment_dict, adverb_dict):
    """
    Boson_Model Adjust factors
    :param data_dict: {'uuid':['text', '0 or 1 or 2']}
    :param Boson_dict: {'Words':'Boson_Score'}
    :param compile_sentiment_dict: {'compile phrase':score}
    :param sentiment_dict: {'pos':[...],'neg':[...]}
    :param adverb_dict: {'extreme':[...],...,'insufficently':[...]}
    :return: return Boson method result
    """
    Boson_Model_Result_dict = {}
    stop_word_list = get_stop_word()
    privative_list = get_privative_list()
    for uuid in data_dict:
        Boson_Model_Result_dict[uuid] = get_Boson_score(data_dict[uuid][0], Boson_dict, stop_word_list,compile_sentiment_dict, sentiment_dict,adverb_dict,privative_list)
    for uuid in Boson_Model_Result_dict:
        score = float(Boson_Model_Result_dict[uuid])
        # print (score)
        if score < lower:
            Boson_Model_Result_dict[uuid] = 0
        if score > upper:
            Boson_Model_Result_dict[uuid] = 2
        if score >= lower and score <= upper:
            Boson_Model_Result_dict[uuid] = 1
    Ratio, pos_ratio, neu_ratio, neg_ratio = Accuracy_Boson_Model(data_dict, Boson_Model_Result_dict)
    return Ratio, pos_ratio, neu_ratio, neg_ratio,Boson_Model_Result_dict

def Boson_Dict_Method():
    """
    :return: several accuracy ratio score
    """
    start = time.time()
    data_dict = load_data()
    # Boson_dict = {'Words':'Boson_Score'}
    boson_dict = Boson_dict()
    compile_sentiment_dict, sentiment_dict = get_sentiment_dict(boson_dict)
    adverb_dict = get_adverb_dict()
    # add into jieba dict
    load_jieba(compile_sentiment_dict)
    ratio, ratio_pos, ratio_neu, ratio_neg, Boson_Model_Result_dict = boson_Dict_Model(data_dict, boson_dict, compile_sentiment_dict, sentiment_dict, adverb_dict)
    print('Bonson_Model_Ratio: ', ratio)
    print('Pos_Ratio: ', ratio_pos)
    print('Neu_Ratio: ', ratio_neu)
    print('Neg_Model_Ratio: ', ratio_neg)
    end = time.time()
    print("Total Time: ", end - start)
    return ratio

if __name__ == "__main__":
    globals()
    upper_list = [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4]
    lower_list = [-0.5, -1, -1.5, -2, -2.5, -3, -3.5, -4]
    ratio_list = []
    for i in tqdm(range(len(upper_list))):
        upper = upper_list[i]
        for j in range(len(upper_list)):
            lower = lower_list[j]
            ratio = Boson_Dict_Method()
            ratio_list.append(ratio)

    # #对比每个upper的不同lower的accuracy提升图

    # fig = plt.figure(figsize = (16,4))
    # plt.subplot(181)
    # x = lower_list
    # y = ratio_list[:8]
    # plt.plot(x,y,color = 'r', linestyle = '-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("0.5 upper")
    #
    # plt.subplot(182)
    # x = lower_list
    # y = ratio_list[8:16]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("1.0 upper")
    #
    # plt.subplot(183)
    # x = lower_list
    # y = ratio_list[16:24]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("1.5 upper")
    #
    # plt.subplot(184)
    # x = lower_list
    # y = ratio_list[24:32]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("2.0 upper")
    #
    # plt.subplot(185)
    # x = lower_list
    # y = ratio_list[32:40]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("2.5 upper")
    #
    # plt.subplot(186)
    # x = lower_list
    # y = ratio_list[40:48]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("3.0 upper")
    #
    # plt.subplot(187)
    # x = lower_list
    # y = ratio_list[48:56]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("3.5 upper")
    #
    # plt.subplot(188)
    # x = lower_list
    # y = ratio_list[56:64]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("4.0 upper")
    # plt.show()




